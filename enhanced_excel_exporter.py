#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
增强的Excel导出器
支持多sheet、数据格式化、图表生成和大数据分批导出
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
from pathlib import Path
from datetime import datetime, date, timedelta
from typing import Dict, List, Optional, Union, Any
from loguru import logger
from enhanced_database import enhanced_db_manager
from config import config


class EnhancedExcelExporter:
    """增强的Excel导出器"""

    def __init__(self):
        self.output_dir = Path('./excel_exports')
        self.output_dir.mkdir(exist_ok=True)

        # 定义样式
        self.setup_styles()

    def setup_styles(self):
        """设置Excel样式"""
        # 标题样式
        self.title_style = NamedStyle(name="title_style")
        self.title_style.font = Font(size=14, bold=True, color="FFFFFF")
        self.title_style.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.title_style.alignment = Alignment(horizontal="center", vertical="center")

        # 表头样式
        self.header_style = NamedStyle(name="header_style")
        self.header_style.font = Font(size=11, bold=True, color="FFFFFF")
        self.header_style.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_style.alignment = Alignment(horizontal="center", vertical="center")
        self.header_style.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # 数据样式
        self.data_style = NamedStyle(name="data_style")
        self.data_style.font = Font(size=10)
        self.data_style.alignment = Alignment(horizontal="center", vertical="center")
        self.data_style.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

    def export_all_stock_data(self,
                              include_basic_data: bool = True,
                              include_tick_data: bool = False,
                              include_indicators: bool = True,
                              recent_days: int = 30) -> str:
        """导出所有股票数据到一个Excel文件"""

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = self.output_dir / f"全部股票数据_{timestamp}.xlsx"

        try:
            logger.info(f"开始导出所有股票数据到: {filename}")

            wb = Workbook()

            # 删除默认工作表
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # 1. 导出股票列表
            self._export_stock_list(wb)

            # 2. 导出股票基本信息
            self._export_stock_info(wb)

            # 3. 导出最新交易数据摘要
            if include_basic_data:
                self._export_latest_trading_summary(wb, recent_days)

            # 4. 导出热门股票详细数据
            if include_basic_data:
                self._export_popular_stocks_detail(wb, recent_days)

            # 5. 导出技术指标摘要
            if include_indicators:
                self._export_indicators_summary(wb)

            # 6. 导出市场统计
            self._export_market_statistics(wb)

            # 7. 如果包含分笔数据，导出最新分笔数据样本
            if include_tick_data:
                self._export_tick_data_sample(wb)

            # 保存文件
            wb.save(filename)

            logger.info(f"Excel导出完成: {filename}")
            return str(filename)

        except Exception as e:
            logger.error(f"导出Excel失败: {e}")
            raise

    def _export_stock_list(self, wb: Workbook):
        """导出股票列表"""
        try:
            ws = wb.create_sheet("股票列表")

            # 获取股票列表数据
            sql = """
            SELECT stock_code, stock_name, market, industry, list_date,
                   total_shares, float_shares
            FROM stock_info
            ORDER BY market, stock_code
            """

            df = enhanced_db_manager.query_to_dataframe(sql)

            if not df.empty:
                # 添加标题
                ws['A1'] = "A股股票列表"
                ws.merge_cells('A1:G1')
                ws['A1'].style = self.title_style

                # 添加数据
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)

                # 应用样式
                self._apply_table_style(ws, len(df) + 1, 7, start_row=2)

                # 自动调整列宽
                self._auto_fit_columns(ws)

                logger.info(f"股票列表导出完成: {len(df)} 条记录")
            else:
                ws['A1'] = "无股票数据"

        except Exception as e:
            logger.error(f"导出股票列表失败: {e}")

    def _export_stock_info(self, wb: Workbook):
        """导出股票基本信息"""
        try:
            ws = wb.create_sheet("股票基本信息")

            sql = """
            SELECT stock_code, stock_name, market, industry, list_date,
                   total_shares, float_shares, updated_at
            FROM stock_info
            ORDER BY market, stock_code
            """

            df = enhanced_db_manager.query_to_dataframe(sql)

            if not df.empty:
                # 数据处理
                if 'total_shares' in df.columns:
                    df['总股本(万股)'] = (df['total_shares'] / 10000).round(2)
                if 'float_shares' in df.columns:
                    df['流通股本(万股)'] = (df['float_shares'] / 10000).round(2)

                # 重命名列
                df = df.rename(columns={
                    'stock_code': '股票代码',
                    'stock_name': '股票名称',
                    'market': '市场',
                    'industry': '行业',
                    'list_date': '上市日期',
                    'updated_at': '更新时间'
                })

                # 选择要导出的列
                export_cols = ['股票代码', '股票名称', '市场', '行业', '上市日期',
                               '总股本(万股)', '流通股本(万股)', '更新时间']
                df_export = df[export_cols]

                # 添加标题
                ws['A1'] = "股票基本信息详表"
                ws.merge_cells('A1:H1')
                ws['A1'].style = self.title_style

                # 添加数据
                for r in dataframe_to_rows(df_export, index=False, header=True):
                    ws.append(r)

                # 应用样式
                self._apply_table_style(ws, len(df_export) + 1, len(export_cols), start_row=2)

                # 自动调整列宽
                self._auto_fit_columns(ws)

                logger.info(f"股票基本信息导出完成: {len(df_export)} 条记录")

        except Exception as e:
            logger.error(f"导出股票基本信息失败: {e}")

    def _export_latest_trading_summary(self, wb: Workbook, recent_days: int = 30):
        """导出最新交易数据摘要"""
        try:
            ws = wb.create_sheet("最新交易摘要")

            # 获取最新交易日期
            date_sql = "SELECT MAX(trade_date) as latest_date FROM basic_data WHERE period = 'daily'"
            date_result = enhanced_db_manager.query_to_dataframe(date_sql)

            if date_result.empty or date_result.iloc[0]['latest_date'] is None:
                ws['A1'] = "无交易数据"
                return

            latest_date = date_result.iloc[0]['latest_date']

            # 获取最新交易数据
            sql = f"""
            SELECT b.stock_code, s.stock_name, s.market, s.industry,
                   b.close_price, b.volume, b.amount, b.turnover_rate,
                   ROUND((b.close_price - LAG(b.close_price) OVER (PARTITION BY b.stock_code ORDER BY b.trade_date)) /
                         LAG(b.close_price) OVER (PARTITION BY b.stock_code ORDER BY b.trade_date) * 100, 2) as price_change_pct
            FROM basic_data b
            LEFT JOIN stock_info s ON b.stock_code = s.stock_code
            WHERE b.period = 'daily'
              AND b.trade_date = '{latest_date}'
              AND b.close_price > 0
            ORDER BY b.amount DESC
            LIMIT 1000
            """

            df = enhanced_db_manager.query_to_dataframe(sql)

            if not df.empty:
                # 数据处理
                df['成交量(万股)'] = (df['volume'] / 10000).round(2)
                df['成交额(万元)'] = (df['amount'] / 10000).round(2)

                # 重命名列
                df = df.rename(columns={
                    'stock_code': '股票代码',
                    'stock_name': '股票名称',
                    'market': '市场',
                    'industry': '行业',
                    'close_price': '收盘价',
                    'turnover_rate': '换手率(%)',
                    'price_change_pct': '涨跌幅(%)'
                })

                export_cols = ['股票代码', '股票名称', '市场', '行业', '收盘价',
                               '成交量(万股)', '成交额(万元)', '换手率(%)', '涨跌幅(%)']
                df_export = df[export_cols]

                # 添加标题
                ws['A1'] = f"最新交易数据摘要 ({latest_date})"
                ws.merge_cells('A1:I1')
                ws['A1'].style = self.title_style

                # 添加数据
                for r in dataframe_to_rows(df_export, index=False, header=True):
                    ws.append(r)

                # 应用样式
                self._apply_table_style(ws, len(df_export) + 1, len(export_cols), start_row=2)

                # 为涨跌幅列添加条件格式
                self._add_conditional_formatting(ws, 'I', len(df_export) + 2)

                # 自动调整列宽
                self._auto_fit_columns(ws)

                logger.info(f"最新交易摘要导出完成: {len(df_export)} 条记录")

        except Exception as e:
            logger.error(f"导出最新交易摘要失败: {e}")

    def _export_popular_stocks_detail(self, wb: Workbook, recent_days: int = 30):
        """导出热门股票详细数据"""
        try:
            ws = wb.create_sheet("热门股票详情")

            # 获取热门股票（按成交额排序，取前50只）
            sql = f"""
            SELECT stock_code, AVG(amount) as avg_amount
            FROM basic_data
            WHERE period = 'daily'
              AND trade_date >= DATE_SUB(CURDATE(), INTERVAL {recent_days} DAY)
            GROUP BY stock_code
            ORDER BY avg_amount DESC
            LIMIT 50
            """

            popular_stocks = enhanced_db_manager.query_to_dataframe(sql)

            if not popular_stocks.empty:
                stock_codes = "', '".join(popular_stocks['stock_code'].tolist())

                # 获取这些股票的详细数据
                detail_sql = f"""
                SELECT b.stock_code, s.stock_name, b.trade_date,
                       b.open_price, b.high_price, b.low_price, b.close_price,
                       b.volume, b.amount, b.turnover_rate
                FROM basic_data b
                LEFT JOIN stock_info s ON b.stock_code = s.stock_code
                WHERE b.period = 'daily'
                  AND b.stock_code IN ('{stock_codes}')
                  AND b.trade_date >= DATE_SUB(CURDATE(), INTERVAL {recent_days} DAY)
                ORDER BY b.stock_code, b.trade_date DESC
                """

                df = enhanced_db_manager.query_to_dataframe(detail_sql)

                if not df.empty:
                    # 数据处理
                    df['成交量(万股)'] = (df['volume'] / 10000).round(2)
                    df['成交额(万元)'] = (df['amount'] / 10000).round(2)

                    # 重命名列
                    df = df.rename(columns={
                        'stock_code': '股票代码',
                        'stock_name': '股票名称',
                        'trade_date': '交易日期',
                        'open_price': '开盘价',
                        'high_price': '最高价',
                        'low_price': '最低价',
                        'close_price': '收盘价',
                        'turnover_rate': '换手率(%)'
                    })

                    export_cols = ['股票代码', '股票名称', '交易日期', '开盘价', '最高价',
                                   '最低价', '收盘价', '成交量(万股)', '成交额(万元)', '换手率(%)']
                    df_export = df[export_cols]

                    # 添加标题
                    ws['A1'] = f"热门股票详细数据 (最近{recent_days}天)"
                    ws.merge_cells('A1:J1')
                    ws['A1'].style = self.title_style

                    # 添加数据
                    for r in dataframe_to_rows(df_export, index=False, header=True):
                        ws.append(r)

                    # 应用样式
                    self._apply_table_style(ws, len(df_export) + 1, len(export_cols), start_row=2)

                    # 自动调整列宽
                    self._auto_fit_columns(ws)

                    logger.info(f"热门股票详情导出完成: {len(df_export)} 条记录")

        except Exception as e:
            logger.error(f"导出热门股票详情失败: {e}")

    def _export_indicators_summary(self, wb: Workbook):
        """导出技术指标摘要"""
        try:
            ws = wb.create_sheet("技术指标摘要")

            # 获取技术指标数据
            sql = """
            SELECT i.stock_code, s.stock_name, i.indicator_name,
                   i.indicator_value, i.trade_date
            FROM indicator_data i
            LEFT JOIN stock_info s ON i.stock_code = s.stock_code
            WHERE i.trade_date = (
                SELECT MAX(trade_date) FROM indicator_data
                WHERE stock_code = i.stock_code AND indicator_name = i.indicator_name
            )
            ORDER BY i.stock_code, i.indicator_name
            """

            df = enhanced_db_manager.query_to_dataframe(sql)

            if not df.empty:
                # 数据透视，将指标名称作为列
                pivot_df = df.pivot_table(
                    index=['stock_code', 'stock_name'],
                    columns='indicator_name',
                    values='indicator_value',
                    aggfunc='first'
                ).reset_index()

                # 重命名列
                pivot_df = pivot_df.rename(columns={
                    'stock_code': '股票代码',
                    'stock_name': '股票名称'
                })

                # 添加标题
                ws['A1'] = "技术指标摘要 (最新数据)"
                col_count = len(pivot_df.columns)
                ws.merge_cells(f'A1:{chr(64 + col_count)}1')
                ws['A1'].style = self.title_style

                # 添加数据
                for r in dataframe_to_rows(pivot_df, index=False, header=True):
                    ws.append(r)

                # 应用样式
                self._apply_table_style(ws, len(pivot_df) + 1, col_count, start_row=2)

                # 自动调整列宽
                self._auto_fit_columns(ws)

                logger.info(f"技术指标摘要导出完成: {len(pivot_df)} 条记录")
            else:
                ws['A1'] = "无技术指标数据"

        except Exception as e:
            logger.error(f"导出技术指标摘要失败: {e}")

    def _get_tick_data_statistics(self):
        """获取分笔数据统计信息"""
        try:
            # 查找所有tick_data表
            tables_sql = """
            SHOW TABLES LIKE 'tick_data_%'
            """
            tables_result = enhanced_db_manager.query_to_dataframe(tables_sql)

            if tables_result.empty:
                return {'stocks_with_tick_data': 0, 'latest_tick_date': None}

            # 找到最新的表
            table_names = []
            for _, row in tables_result.iterrows():
                table_name = list(row.values)[0]  # 获取表名
                if 'tick_data_' in table_name:
                    table_names.append(table_name)

            if not table_names:
                return {'stocks_with_tick_data': 0, 'latest_tick_date': None}

            # 按日期排序，获取最新的表
            table_names.sort(reverse=True)
            latest_table = table_names[0]

            # 获取统计信息
            stats_sql = f"""
            SELECT
                COUNT(DISTINCT stock_code) as stocks_with_tick_data,
                MAX(trade_date) as latest_tick_date
            FROM {latest_table}
            """

            result = enhanced_db_manager.query_to_dataframe(stats_sql)
            if not result.empty:
                return {
                    'stocks_with_tick_data': result.iloc[0]['stocks_with_tick_data'],
                    'latest_tick_date': result.iloc[0]['latest_tick_date']
                }
            else:
                return {'stocks_with_tick_data': 0, 'latest_tick_date': None}

        except Exception as e:
            logger.warning(f"获取分笔数据统计失败: {e}")
            return {'stocks_with_tick_data': 0, 'latest_tick_date': None}

    def _export_market_statistics(self, wb: Workbook):
        """导出市场统计数据"""
        try:
            ws = wb.create_sheet("市场统计")

            # 获取各种统计数据
            stats = []

            # 1. 股票总数统计
            stock_count_sql = """
            SELECT
                COUNT(*) as total_stocks,
                SUM(CASE WHEN market = 'sh' THEN 1 ELSE 0 END) as sh_stocks,
                SUM(CASE WHEN market = 'sz' THEN 1 ELSE 0 END) as sz_stocks
            FROM stock_info
            """
            stock_stats = enhanced_db_manager.query_to_dataframe(stock_count_sql)

            if not stock_stats.empty:
                stats.append(['股票总数', stock_stats.iloc[0]['total_stocks']])
                stats.append(['上海A股', stock_stats.iloc[0]['sh_stocks']])
                stats.append(['深圳A股', stock_stats.iloc[0]['sz_stocks']])

            # 2. 行业分布统计
            industry_sql = """
            SELECT industry, COUNT(*) as count
            FROM stock_info
            WHERE industry IS NOT NULL AND industry != ''
            GROUP BY industry
            ORDER BY count DESC
            LIMIT 10
            """
            industry_stats = enhanced_db_manager.query_to_dataframe(industry_sql)

            # 3. 数据统计 - 使用安全查询方式
            data_stats = pd.DataFrame()
            stats_data = {}

            # 检查basic_data表
            if enhanced_db_manager.table_exists('basic_data'):
                try:
                    basic_stats_sql = """
                    SELECT
                        COUNT(DISTINCT stock_code) as stocks_with_basic_data,
                        MAX(trade_date) as latest_basic_date
                    FROM basic_data
                    WHERE period = 'daily'
                    """
                    basic_result = enhanced_db_manager.query_to_dataframe(basic_stats_sql)
                    if not basic_result.empty:
                        stats_data['stocks_with_basic_data'] = basic_result.iloc[0]['stocks_with_basic_data']
                        stats_data['latest_basic_date'] = basic_result.iloc[0]['latest_basic_date']
                except Exception as e:
                    logger.warning(f"查询basic_data统计失败: {e}")
                    stats_data['stocks_with_basic_data'] = 0
                    stats_data['latest_basic_date'] = None
            else:
                stats_data['stocks_with_basic_data'] = 0
                stats_data['latest_basic_date'] = None

            # 检查tick_data表 - 由于是分表结构，需要查找最新的表
            tick_stats_data = self._get_tick_data_statistics()
            stats_data.update(tick_stats_data)

            # 创建DataFrame
            if stats_data:
                data_stats = pd.DataFrame([stats_data])

            if not data_stats.empty:
                stats.append(['有基础数据的股票数', data_stats.iloc[0]['stocks_with_basic_data']])
                stats.append(['有分笔数据的股票数', data_stats.iloc[0]['stocks_with_tick_data']])
                stats.append(['最新基础数据日期', str(data_stats.iloc[0]['latest_basic_date'])])
                stats.append(['最新分笔数据日期', str(data_stats.iloc[0]['latest_tick_date'])])

            # 创建统计表
            ws['A1'] = "市场统计概览"
            ws.merge_cells('A1:B1')
            ws['A1'].style = self.title_style

            # 基本统计
            ws['A3'] = "指标"
            ws['B3'] = "数值"
            ws['A3'].style = self.header_style
            ws['B3'].style = self.header_style

            row = 4
            for stat in stats:
                ws.cell(row=row, column=1, value=stat[0])
                ws.cell(row=row, column=2, value=stat[1])
                row += 1

            # 行业分布
            if not industry_stats.empty:
                start_row = row + 2
                ws.cell(row=start_row, column=1, value="热门行业分布")
                ws.merge_cells(f'A{start_row}:B{start_row}')
                ws.cell(row=start_row, column=1).style = self.title_style

                ws.cell(row=start_row + 2, column=1, value="行业")
                ws.cell(row=start_row + 2, column=2, value="股票数量")
                ws.cell(row=start_row + 2, column=1).style = self.header_style
                ws.cell(row=start_row + 2, column=2).style = self.header_style

                for idx, row_data in industry_stats.iterrows():
                    ws.cell(row=start_row + 3 + idx, column=1, value=row_data['industry'])
                    ws.cell(row=start_row + 3 + idx, column=2, value=row_data['count'])

            # 自动调整列宽
            self._auto_fit_columns(ws)

            logger.info("市场统计导出完成")

        except Exception as e:
            logger.error(f"导出市场统计失败: {e}")

    def _export_tick_data_sample(self, wb: Workbook):
        """导出分笔数据样本"""
        try:
            ws = wb.create_sheet("分笔数据样本")
            # 获取最新的tick_data表和数据
            tick_stats = self._get_tick_data_statistics()

            if tick_stats['latest_tick_date'] is None:
                ws['A1'] = "暂无分笔数据"
                return

            # 查找最新的表
            tables_sql = "SHOW TABLES LIKE 'tick_data_%'"
            tables_result = enhanced_db_manager.query_to_dataframe(tables_sql)

            if tables_result.empty:
                ws['A1'] = "暂无分笔数据表"
                return

            # 找到最新的表
            table_names = []
            for _, row in tables_result.iterrows():
                table_name = list(row.values)[0]
                if 'tick_data_' in table_name:
                    table_names.append(table_name)

            if not table_names:
                ws['A1'] = "暂无有效的分笔数据表"
                return

            table_names.sort(reverse=True)
            latest_table = table_names[0]

            # 获取最新日期的分笔数据样本
            sql = f"""
            SELECT t.stock_code, s.stock_name, t.trade_time, t.price,
                   t.volume, t.amount, t.trade_type
            FROM {latest_table} t
            LEFT JOIN stock_info s ON t.stock_code = s.stock_code
            WHERE t.trade_date = (
                SELECT MAX(trade_date) FROM {latest_table}
            )
            ORDER BY t.trade_time DESC
            LIMIT 10000
            """
            df = enhanced_db_manager.query_to_dataframe(sql)
            if not df.empty:
                # 重命名列
                df = df.rename(columns={
                    'stock_code': '股票代码',
                    'stock_name': '股票名称',
                    'trade_time': '交易时间',
                    'price': '成交价',
                    'volume': '成交量',
                    'amount': '成交额',
                    'trade_type': '交易类型'
                })
                # 添加标题
                ws['A1'] = f"分笔数据样本 (来自{latest_table}, 最新交易日前10000条)"
                ws.merge_cells('A1:G1')
                ws['A1'].style = self.title_style
                # 添加数据
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)
                # 应用样式
                self._apply_table_style(ws, len(df) + 1, 7, start_row=2)
                # 自动调整列宽
                self._auto_fit_columns(ws)
                logger.info(f"分笔数据样本导出完成: {len(df)} 条记录")
            else:
                ws['A1'] = "暂无分笔数据"

        except Exception as e:
            logger.error(f"导出分笔数据样本失败: {e}")

    def _apply_table_style(self, ws, row_count: int, col_count: int, start_row: int = 1):
        """应用表格样式"""
        try:
            # 应用表头样式
            for col in range(1, col_count + 1):
                cell = ws.cell(row=start_row, column=col)
                cell.style = self.header_style

            # 应用数据样式
            for row in range(start_row + 1, row_count + start_row):
                for col in range(1, col_count + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.style = self.data_style

        except Exception as e:
            logger.error(f"应用表格样式失败: {e}")

    def _auto_fit_columns(self, ws):
        """自动调整列宽"""
        try:
            for column in ws.columns:
                max_length = 0
                column_letter = None

                for cell in column:
                    # 跳过合并单元格
                    if hasattr(cell, 'coordinate') and cell.coordinate in ws.merged_cells:
                        continue

                    # 获取列字母
                    if column_letter is None:
                        if hasattr(cell, 'column_letter'):
                            column_letter = cell.column_letter
                        else:
                            # 处理MergedCell对象没有column_letter属性的情况
                            from openpyxl.utils import get_column_letter
                            column_letter = get_column_letter(cell.column)

                    try:
                        cell_value = str(cell.value) if cell.value is not None else ""
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                    except:
                        pass

                if column_letter:
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

        except Exception as e:
            logger.error(f"自动调整列宽失败: {e}")

    def _add_conditional_formatting(self, ws, column: str, row_count: int):
        """添加条件格式"""
        try:
            # 为涨跌幅列添加颜色条件格式
            red_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            green_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")

            range_address = f"{column}3:{column}{row_count}"

            # 添加数据条
            data_bar = DataBarRule(
                start_type='num', start_value=-10,
                end_type='num', end_value=10,
                color="4CAF50"
            )
            ws.conditional_formatting.add(range_address, data_bar)

        except Exception as e:
            logger.error(f"添加条件格式失败: {e}")

    def export_stock_detail_by_code(self, stock_code: str, days: int = 90) -> str:
        """导出单个股票的详细数据"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = self.output_dir / f"股票_{stock_code}_详细数据_{timestamp}.xlsx"

            wb = Workbook()
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # 获取股票基本信息
            stock_info_sql = "SELECT * FROM stock_info WHERE stock_code = :stock_code"
            stock_info = enhanced_db_manager.query_to_dataframe(stock_info_sql, {'stock_code': stock_code})

            if stock_info.empty:
                raise ValueError(f"股票代码 {stock_code} 不存在")

            stock_name = stock_info.iloc[0]['stock_name']

            # 1. 基本信息sheet
            info_ws = wb.create_sheet(f"{stock_name}_基本信息")
            for r in dataframe_to_rows(stock_info, index=False, header=True):
                info_ws.append(r)
            self._auto_fit_columns(info_ws)

            # 2. K线数据sheet
            basic_sql = f"""
            SELECT * FROM basic_data
            WHERE stock_code = :stock_code AND period = 'daily'
              AND trade_date >= DATE_SUB(CURDATE(), INTERVAL {days} DAY)
            ORDER BY trade_date DESC
            """
            basic_data = enhanced_db_manager.query_to_dataframe(basic_sql, {'stock_code': stock_code})

            if not basic_data.empty:
                basic_ws = wb.create_sheet(f"{stock_name}_K线数据")
                for r in dataframe_to_rows(basic_data, index=False, header=True):
                    basic_ws.append(r)
                self._auto_fit_columns(basic_ws)

            # 3. 技术指标sheet
            indicator_sql = f"""
            SELECT * FROM indicator_data
            WHERE stock_code = :stock_code
              AND trade_date >= DATE_SUB(CURDATE(), INTERVAL {days} DAY)
            ORDER BY trade_date DESC, indicator_name
            """
            indicator_data = enhanced_db_manager.query_to_dataframe(indicator_sql, {'stock_code': stock_code})

            if not indicator_data.empty:
                indicator_ws = wb.create_sheet(f"{stock_name}_技术指标")
                for r in dataframe_to_rows(indicator_data, index=False, header=True):
                    indicator_ws.append(r)
                self._auto_fit_columns(indicator_ws)

            wb.save(filename)
            logger.info(f"股票 {stock_code} 详细数据导出完成: {filename}")
            return str(filename)

        except Exception as e:
            logger.error(f"导出股票 {stock_code} 详细数据失败: {e}")
            raise


# 创建全局实例
enhanced_excel_exporter = EnhancedExcelExporter()
