#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel导出器补丁 - 添加新的导出方法
"""

import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
from loguru import logger
from data.enhanced_database import enhanced_db_manager


def export_basic_data_by_period(self, period: str, limit: int = 10000) -> str:
    """按时间周期导出基础数据"""
    try:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = self.output_dir / f'basic_data_{period}_{timestamp}.xlsx'

        # 创建工作簿
        wb = Workbook()
        wb.remove(wb.active)  # 删除默认sheet

        # 设置样式
        self.setup_styles()

        # 查询基础数据
        table_name = f"basic_data_{period}"

        sql = f"""
        SELECT 
            stock_code, trade_date, open_price, high_price, 
            low_price, close_price, volume, amount, turnover_rate,
            price_change, pct_change
        FROM {table_name}
        ORDER BY trade_date DESC, stock_code
        LIMIT {limit}
        """

        data = enhanced_db_manager.safe_query_to_dataframe(
            sql,
            required_tables=[table_name]
        )

        if data.empty:
            print(f"❌ {period} 周期没有数据")
            return ""

        # 创建主数据工作表
        ws = wb.create_sheet(f'{period.upper()}数据')

        # 写入数据
        for r in dataframe_to_rows(data, index=False, header=True):
            ws.append(r)

        # 应用格式化
        self._auto_fit_columns(ws)
        self._apply_table_style(ws, len(data) + 1, len(data.columns))

        # 添加条件格式化
        if 'pct_change' in data.columns:
            self._add_conditional_formatting(ws, 'K', len(data) + 1)  # 涨跌幅列

        # 创建统计汇总表
        summary_ws = wb.create_sheet('数据统计')

        # 基本统计信息
        stats = [
            ['统计项', '数值'],
            ['数据条数', len(data)],
            ['股票数量', data['stock_code'].nunique() if not data.empty else 0],
            ['数据周期', period],
            ['最新交易日', data['trade_date'].max() if not data.empty else '无'],
            ['最早交易日', data['trade_date'].min() if not data.empty else '无'],
        ]

        if 'pct_change' in data.columns:
            stats.extend([
                ['平均涨跌幅(%)', round(data['pct_change'].mean(), 2)],
                ['最大涨幅(%)', round(data['pct_change'].max(), 2)],
                ['最大跌幅(%)', round(data['pct_change'].min(), 2)],
            ])

        for row in stats:
            summary_ws.append(row)

        self._auto_fit_columns(summary_ws)

        # 保存文件
        wb.save(filename)
        logger.info(f"{period} 周期基础数据导出完成: {filename}")
        return str(filename)

    except Exception as e:
        logger.error(f"导出 {period} 周期基础数据失败: {e}")
        raise


def export_analysis_results(self, analysis_results: dict) -> str:
    """导出所有分析结果到Excel"""
    try:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = self.output_dir / f'analysis_results_{timestamp}.xlsx'

        # 创建工作簿
        wb = Workbook()
        wb.remove(wb.active)  # 删除默认sheet

        # 设置样式
        self.setup_styles()

        # 导出三层共振分析结果
        if analysis_results.get('resonance'):
            ws = wb.create_sheet('三层共振分析')
            resonance_data = pd.DataFrame(analysis_results['resonance'])

            for r in dataframe_to_rows(resonance_data, index=False, header=True):
                ws.append(r)

            self._auto_fit_columns(ws)
            self._apply_table_style(ws, len(resonance_data) + 1, len(resonance_data.columns))

        # 导出涨停板分析结果
        if analysis_results.get('limit_up'):
            ws = wb.create_sheet('涨停板分析')
            limit_up_data = pd.DataFrame(analysis_results['limit_up'])

            for r in dataframe_to_rows(limit_up_data, index=False, header=True):
                ws.append(r)

            self._auto_fit_columns(ws)
            self._apply_table_style(ws, len(limit_up_data) + 1, len(limit_up_data.columns))

        # 导出异动检测结果
        if analysis_results.get('anomaly'):
            ws = wb.create_sheet('异动检测')
            anomaly_data = pd.DataFrame(analysis_results['anomaly'])

            for r in dataframe_to_rows(anomaly_data, index=False, header=True):
                ws.append(r)

            self._auto_fit_columns(ws)
            self._apply_table_style(ws, len(anomaly_data) + 1, len(anomaly_data.columns))

        # 导出多空通道分析结果
        if analysis_results.get('channel'):
            ws = wb.create_sheet('多空通道分析')
            channel_data = pd.DataFrame(analysis_results['channel'])

            for r in dataframe_to_rows(channel_data, index=False, header=True):
                ws.append(r)

            self._auto_fit_columns(ws)
            self._apply_table_style(ws, len(channel_data) + 1, len(channel_data.columns))

        # 创建分析汇总表
        summary_ws = wb.create_sheet('分析汇总')

        summary_data = [
            ['分析类型', '结果数量', '分析状态'],
            ['三层共振分析', len(analysis_results.get('resonance', [])),
             '✅ 完成' if analysis_results.get('resonance') else '❌ 无数据'],
            ['涨停板分析', len(analysis_results.get('limit_up', [])),
             '✅ 完成' if analysis_results.get('limit_up') else '❌ 无数据'],
            ['异动检测', len(analysis_results.get('anomaly', [])),
             '✅ 完成' if analysis_results.get('anomaly') else '❌ 无数据'],
            ['多空通道分析', len(analysis_results.get('channel', [])),
             '✅ 完成' if analysis_results.get('channel') else '❌ 无数据'],
            ['', '', ''],
            ['导出时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), ''],
            ['总分析项目', '4', ''],
            ['成功项目', sum(1 for k, v in analysis_results.items() if v), ''],
        ]

        for row in summary_data:
            summary_ws.append(row)

        self._auto_fit_columns(summary_ws)

        # 保存文件
        wb.save(filename)
        logger.info(f"分析结果导出完成: {filename}")
        return str(filename)

    except Exception as e:
        logger.error(f"导出分析结果失败: {e}")
        raise


# 动态添加方法到EnhancedExcelExporter类
def patch_excel_exporter():
    """为Excel导出器添加新方法"""
    try:
        from export.enhanced_excel_exporter import EnhancedExcelExporter, enhanced_excel_exporter

        # 添加新方法到类
        EnhancedExcelExporter.export_basic_data_by_period = export_basic_data_by_period
        EnhancedExcelExporter.export_analysis_results = export_analysis_results

        logger.info("Excel导出器补丁应用成功")
        return True

    except Exception as e:
        logger.error(f"应用Excel导出器补丁失败: {e}")
        return False