# -*- coding: utf-8 -*-
"""
增强的Excel导出器
支持多sheet、数据格式化、图表生成和大数据分批导出
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
from pathlib import Path
from datetime import datetime, date, timedelta
from typing import Dict, List, Optional, Union, Any
from loguru import logger
from data.enhanced_database import enhanced_db_manager
from core.config import config


class EnhancedExcelExporter:
    """增强的Excel导出器"""

    def __init__(self):
        self.output_dir = Path('./excel_exports')
        self.output_dir.mkdir(exist_ok=True)

        # 定义样式
        self.setup_styles()

    def setup_styles(self):
        """设置Excel样式"""
        # 标题样式
        self.title_style = NamedStyle(name="title_style")
        self.title_style.font = Font(size=14, bold=True, color="FFFFFF")
        self.title_style.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.title_style.alignment = Alignment(horizontal="center", vertical="center")

        # 表头样式
        self.header_style = NamedStyle(name="header_style")
        self.header_style.font = Font(size=11, bold=True, color="FFFFFF")
        self.header_style.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_style.alignment = Alignment(horizontal="center", vertical="center")
        self.header_style.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # 数据样式
        self.data_style = NamedStyle(name="data_style")
        self.data_style.font = Font(size=10)
        self.data_style.alignment = Alignment(horizontal="center", vertical="center")
        self.data_style.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

    def export_all_stock_data(self,
                              include_basic_data: bool = True,
                              include_tick_data: bool = False,
                              include_indicators: bool = True,
                              recent_days: int = 30) -> str:
        """导出所有股票数据到一个Excel文件"""

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = self.output_dir / f"全部股票数据_{timestamp}.xlsx"

        try:
            logger.info(f"开始导出所有股票数据到: {filename}")

            wb = Workbook()

            # 删除默认工作表
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # 1. 导出股票列表
            self._export_stock_list(wb)

            # 2. 导出股票基本信息
            self._export_stock_info(wb)

            # 3. 导出最新交易数据摘要
            if include_basic_data:
                self._export_latest_trading_summary(wb, recent_days)

            # 4. 导出热门股票详细数据
            if include_basic_data:
                self._export_popular_stocks_detail(wb, recent_days)

            # 5. 导出技术指标摘要
            if include_indicators:
                self._export_indicators_summary(wb)

            # 6. 导出市场统计
            self._export_market_statistics(wb)

            # 7. 如果包含分笔数据，导出最新分笔数据样本
            if include_tick_data:
                self._export_tick_data_sample(wb)

            # 保存文件
            wb.save(filename)

            logger.info(f"Excel导出完成: {filename}")
            return str(filename)

        except Exception as e:
            logger.error(f"导出Excel失败: {e}")
            raise

    def _export_stock_list(self, wb: Workbook):
        """导出股票列表"""
        try:
            ws = wb.create_sheet("股票列表")

            # 获取股票列表数据
            sql = """
            SELECT stock_code, stock_name, market, industry, list_date,
                   total_shares, float_shares
            FROM stock_info
            ORDER BY market, stock_code
            """

            df = enhanced_db_manager.query_to_dataframe(sql)

            if not df.empty:
                # 添加标题
                ws['A1'] = "A股股票列表"
                ws.merge_cells('A1:G1')
                ws['A1'].style = self.title_style

                # 添加数据
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)

                # 应用样式
                self._apply_table_style(ws, len(df) + 1, 7, start_row=2)

                # 自动调整列宽
                self._auto_fit_columns(ws)

                logger.info(f"股票列表导出完成: {len(df)} 条记录")
            else:
                ws['A1'] = "无股票数据"

        except Exception as e:
            logger.error(f"导出股票列表失败: {e}")

    def _export_stock_info(self, wb: Workbook):
        """导出股票基本信息"""
        try:
            ws = wb.create_sheet("股票基本信息")

            sql = """
            SELECT stock_code, stock_name, market, industry, list_date,
                   total_shares, float_shares, updated_at
            FROM stock_info
            ORDER BY market, stock_code
            """

            df = enhanced_db_manager.query_to_dataframe(sql)

            if not df.empty:
                # 数据处理
                if 'total_shares' in df.columns:
                    df['总股本(万股)'] = (df['total_shares'] / 10000).round(2)
                if 'float_shares' in df.columns:
                    df['流通股本(万股)'] = (df['float_shares'] / 10000).round(2)

                # 重命名列
                df = df.rename(columns={
                    'stock_code': '股票代码',
                    'stock_name': '股票名称',
                    'market': '市场',
                    'industry': '行业',
                    'list_date': '上市日期',
                    'updated_at': '更新时间'
                })

                # 选择要导出的列
                export_cols = ['股票代码', '股票名称', '市场', '行业', '上市日期',
                               '总股本(万股)', '流通股本(万股)', '更新时间']
                df_export = df[export_cols]

                # 添加标题
                ws['A1'] = "股票基本信息详表"
                ws.merge_cells('A1:H1')
                ws['A1'].style = self.title_style

                # 添加数据
                for r in dataframe_to_rows(df_export, index=False, header=True):
                    ws.append(r)

                # 应用样式
                self._apply_table_style(ws, len(df_export) + 1, len(export_cols), start_row=2)

                # 自动调整列宽
                self._auto_fit_columns(ws)

                logger.info(f"股票基本信息导出完成: {len(df_export)} 条记录")

        except Exception as e:
            logger.error(f"导出股票基本信息失败: {e}")

    def _export_latest_trading_summary(self, wb: Workbook, recent_days: int = 30):
        """导出最新交易数据摘要"""
        try:
            ws = wb.create_sheet("最新交易摘要")

            # 获取最新交易日期
            date_sql = "SELECT MAX(trade_date) as latest_date FROM basic_data WHERE period = 'daily'"
            date_result = enhanced_db_manager.query_to_dataframe(date_sql)

            if date_result.empty or date_result.iloc[0]['latest_date'] is None:
                ws['A1'] = "无交易数据"
                return

            latest_date = date_result.iloc[0]['latest_date']

            # 获取最新交易数据
            sql = f"""
            SELECT b.stock_code, s.stock_name, s.market, s.industry,
                   b.close_price, b.volume, b.amount, b.turnover_rate,
                   ROUND((b.close_price - LAG(b.close_price) OVER (PARTITION BY b.stock_code ORDER BY b.trade_date)) /
                         LAG(b.close_price) OVER (PARTITION BY b.stock_code ORDER BY b.trade_date) * 100, 2) as price_change_pct
            FROM basic_data b
            LEFT JOIN stock_info s ON b.stock_code = s.stock_code
            WHERE b.period = 'daily'
              AND b.trade_date = '{latest_date}'
              AND b.close_price > 0
            ORDER BY b.amount DESC
            LIMIT 1000
            """

            df = enhanced_db_manager.query_to_dataframe(sql)

            if not df.empty:
                # 数据处理
                df['成交量(万股)'] = (df['volume'] / 10000).round(2)
                df['成交额(万元)'] = (df['amount'] / 10000).round(2)

                # 重命名列
                df = df.rename(columns={
                    'stock_code': '股票代码',
                    'stock_name': '股票名称',
                    'market': '市场',
                    'industry': '行业',
                    'close_price': '收盘价',
                    'turnover_rate': '换手率(%)',
                    'price_change_pct': '涨跌幅(%)'
                })

                export_cols = ['股票代码', '股票名称', '市场', '行业', '收盘价',
                               '成交量(万股)', '成交额(万元)', '换手率(%)', '涨跌幅(%)']
                df_export = df[export_cols]

                # 添加标题
                ws['A1'] = f"最新交易数据摘要 ({latest_date})"
                ws.merge_cells('A1:I1')
                ws['A1'].style = self.title_style

                # 添加数据
                for r in dataframe_to_rows(df_export, index=False, header=True):
                    ws.append(r)

                # 应用样式
                self._apply_table_style(ws, len(df_export) + 1, len(export_cols), start_row=2)

                # 为涨跌幅列添加条件格式
                self._add_conditional_formatting(ws, 'I', len(df_export) + 2)

                # 自动调整列宽
                self._auto_fit_columns(ws)

                logger.info(f"最新交易摘要导出完成: {len(df_export)} 条记录")

        except Exception as e:
            logger.error(f"导出最新交易摘要失败: {e}")

    def _export_popular_stocks_detail(self, wb: Workbook, recent_days: int = 30):
        """导出热门股票详细数据"""
        try:
            ws = wb.create_sheet("热门股票详情")

            # 获取热门股票（按成交额排序，取前50只）
            sql = f"""
            SELECT stock_code, AVG(amount) as avg_amount
            FROM basic_data
            WHERE period = 'daily'
              AND trade_date >= DATE_SUB(CURDATE(), INTERVAL {recent_days} DAY)
            GROUP BY stock_code
            ORDER BY avg_amount DESC
            LIMIT 50
            """

            popular_stocks = enhanced_db_manager.query_to_dataframe(sql)

            if not popular_stocks.empty:
                stock_codes = "', '".join(popular_stocks['stock_code'].tolist())

                # 获取这些股票的详细数据
                detail_sql = f"""
                SELECT b.stock_code, s.stock_name, b.trade_date,
                       b.open_price, b.high_price, b.low_price, b.close_price,
                       b.volume, b.amount, b.turnover_rate
                FROM basic_data b
                LEFT JOIN stock_info s ON b.stock_code = s.stock_code
                WHERE b.period = 'daily'
                  AND b.stock_code IN ('{stock_codes}')
                  AND b.trade_date >= DATE_SUB(CURDATE(), INTERVAL {recent_days} DAY)
                ORDER BY b.stock_code, b.trade_date DESC
                """

                df = enhanced_db_manager.query_to_dataframe(detail_sql)

                if not df.empty:
                    # 数据处理
                    df['成交量(万股)'] = (df['volume'] / 10000).round(2)
                    df['成交额(万元)'] = (df['amount'] / 10000).round(2)

                    # 重命名列
                    df = df.rename(columns={
                        'stock_code': '股票代码',
                        'stock_name': '股票名称',
                        'trade_date': '交易日期',
                        'open_price': '开盘价',
                        'high_price': '最高价',
                        'low_price': '最低价',
                        'close_price': '收盘价',
                        'turnover_rate': '换手率(%)'
                    })

                    export_cols = ['股票代码', '股票名称', '交易日期', '开盘价', '最高价',
                                   '最低价', '收盘价', '成交量(万股)', '成交额(万元)', '换手率(%)']
                    df_export = df[export_cols]

                    # 添加标题
                    ws['A1'] = f"热门股票详细数据 (最近{recent_days}天)"
                    ws.merge_cells('A1:J1')
                    ws['A1'].style = self.title_style

                    # 添加数据
                    for r in dataframe_to_rows(df_export, index=False, header=True):
                        ws.append(r)

                    # 应用样式
                    self._apply_table_style(ws, len(df_export) + 1, len(export_cols), start_row=2)

                    # 自动调整列宽
                    self._auto_fit_columns(ws)

                    logger.info(f"热门股票详情导出完成: {len(df_export)} 条记录")

        except Exception as e:
            logger.error(f"导出热门股票详情失败: {e}")

    def _check_table_column_exists(self, table_name, column_name):
        """检查表中是否存在指定字段"""
        try:
            sql = f"SHOW COLUMNS FROM {table_name} LIKE '{column_name}'"
            result = self.db_manager.query_to_dataframe(sql)
            return not result.empty
        except:
            return False

    def _export_indicators_summary(self, wb):
        """导出技术指标摘要"""
        try:
            ws = wb.create_sheet("技术指标摘要")

            # 检查indicator_data表中是否存在必要字段
            has_indicator_value = self._check_table_column_exists('indicator_data', 'indicator_value')
            has_trade_date = self._check_table_column_exists('indicator_data', 'trade_date')

            if not has_indicator_value or not has_trade_date:
                # 如果字段不存在，创建一个提示信息
                ws.append(["技术指标数据", "字段缺失", "请先更新数据库结构"])
                ws.append(["缺失字段", f"indicator_value: {'存在' if has_indicator_value else '缺失'}", f"trade_date: {'存在' if has_trade_date else '缺失'}"])
                logger.warning("indicator_data表缺少必要字段，跳过技术指标摘要导出")
                return

            # 获取技术指标数据
            sql = """
            SELECT i.stock_code, s.stock_name, i.indicator_name,
                   i.indicator_value, i.trade_date
            FROM indicator_data i
            LEFT JOIN stock_info s ON i.stock_code = s.stock_code
            WHERE i.trade_date = (
                SELECT MAX(trade_date) FROM indicator_data
                WHERE stock_code = i.stock_code AND indicator_name = i.indicator_name
            )
            ORDER BY i.stock_code, i.indicator_name
            """

            df = self.db_manager.query_to_dataframe(sql)

            if df.empty:
                ws.append(["股票代码", "股票名称", "指标名称", "指标值", "交易日期"])
                ws.append(["暂无数据", "", "", "", ""])
                logger.info("技术指标数据为空")
                return

            # 数据透视，将指标名称作为列
            pivot_df = df.pivot_table(
                index=['stock_code', 'stock_name'],
                columns='indicator_name',
                values='indicator_value',
                aggfunc='first'
            ).reset_index()

            # 重命名列
            pivot_df = pivot_df.rename(columns={
                'stock_code': '股票代码',
                'stock_name': '股票名称'
            })

            # 添加标题
            ws['A1'] = "技术指标摘要 (最新数据)"
            col_count = len(pivot_df.columns)
            ws.merge_cells(f'A1:{chr(64 + col_count)}1')
            ws['A1'].style = self.title_style

            # 添加数据
            for r in dataframe_to_rows(pivot_df, index=False, header=True):
                ws.append(r)

            # 应用样式
            self._apply_table_style(ws, len(pivot_df) + 1, col_count, start_row=2)

            # 自动调整列宽
            self._auto_fit_columns(ws)

            logger.info(f"技术指标摘要导出完成: {len(pivot_df)} 条记录")
        except Exception as e:
            logger.error(f"导出技术指标摘要失败: {e}")

    def _get_tick_data_statistics(self):
        """获取分笔数据统计信息"""
        try:
            # 查找所有tick_data表
            tables_sql = """
            SHOW TABLES LIKE 'tick_data_%'
            """
            tables_result = enhanced_db_manager.query_to_dataframe(tables_sql)

            if tables_result.empty:
                return {'stocks_with_tick_data': 0, 'latest_tick_date': None}

            # 找到最新的表
            table_names = []
            for _, row in tables_result.iterrows():
                table_name = list(row.values)[0]  # 获取表名
                if 'tick_data_' in table_name:
                    table_names.append(table_name)

            if not table_names:
                return {'stocks_with_tick_data': 0, 'latest_tick_date': None}

            # 按日期排序，获取最新的表
            table_names.sort(reverse=True)
            latest_table = table_names[0]

            # 获取统计信息
            stats_sql = f"""
            SELECT
                COUNT(DISTINCT stock_code) as stocks_with_tick_data,
                MAX(trade_date) as latest_tick_date
            FROM {latest_table}
            """

            result = enhanced_db_manager.query_to_dataframe(stats_sql)
            if not result.empty:
                return {
                    'stocks_with_tick_data': result.iloc[0]['stocks_with_tick_data'],
                    'latest_tick_date': result.iloc[0]['latest_tick_date']
                }
            else:
                return {'stocks_with_tick_data': 0, 'latest_tick_date': None}

        except Exception as e:
            logger.warning(f"获取分笔数据统计失败: {e}")
            return {'stocks_with_tick_data': 0, 'latest_tick_date': None}

    def _export_market_statistics(self, wb):
        """导出市场统计信息"""
        try:
            ws = wb.create_sheet("市场统计")

            # 定义周期表
            periods = ['daily', '1hour', '30min', '15min', '5min', '1min']

            # 查找存在的basic_data表
            available_tables = []
            for period in periods:
                table_name = f"basic_data_{period}"
                has_table = self._check_table_exists(table_name)
                if has_table:
                    available_tables.append((period, table_name))

            if not available_tables:
                ws.append(["市场统计", "无可用数据表", "未找到任何basic_data_{period}表"])
                logger.warning("未找到任何basic_data表")
                return

            # 使用第一个可用的表（通常是daily）进行统计
            primary_period, primary_table = available_tables[0]
            logger.info(f"使用{primary_table}进行市场统计")

            # 检查表中是否有trade_date字段
            has_trade_date = self._check_table_column_exists(primary_table, 'trade_date')

            if not has_trade_date:
                ws.append(["市场统计", "字段缺失", f"{primary_table}表缺少trade_date字段"])
                logger.warning(f"{primary_table}表缺少trade_date字段，使用替代统计方案")

                # 使用替代查询方案
                sql = f"""
                    SELECT
                        COUNT(DISTINCT stock_code) as stocks_with_basic_data,
                        MAX(created_at) as latest_basic_date
                    FROM {primary_table}
                    WHERE period = '{primary_period}'
                """
            else:
                # 原始查询
                sql = f"""
                    SELECT
                        COUNT(DISTINCT stock_code) as stocks_with_basic_data,
                        MAX(trade_date) as latest_basic_date
                    FROM {primary_table}
                    WHERE period = '{primary_period}'
                """

            basic_stats = self.db_manager.query_to_dataframe(sql)

            # 创建统计摘要
            ws.append(["市场统计概览", ""])
            ws.append(["", ""])

            if not basic_stats.empty:
                stats = basic_stats.iloc[0]

                # 写入统计数据
                ws.append(["统计项", "数值"])
                ws.append(["主要数据周期", primary_period])
                ws.append(["有基础数据的股票数", stats.get('stocks_with_basic_data', 0)])
                ws.append(["最新基础数据日期", str(stats.get('latest_basic_date', '未知'))])

                # 统计各周期表的数据
                ws.append(["", ""])
                ws.append(["各周期数据统计", ""])
                ws.append(["周期", "股票数量", "最新日期"])

                for period, table_name in available_tables:
                    try:
                        period_sql = f"""
                        SELECT
                            COUNT(DISTINCT stock_code) as count,
                            MAX(trade_date) as max_date
                        FROM {table_name}
                        WHERE period = '{period}'
                        """
                        period_stats = self.db_manager.query_to_dataframe(period_sql)

                        if not period_stats.empty:
                            period_data = period_stats.iloc[0]
                            ws.append([
                                period,
                                period_data.get('count', 0),
                                str(period_data.get('max_date', '无数据'))
                            ])
                        else:
                            ws.append([period, 0, '无数据'])

                    except Exception as e:
                        logger.warning(f"统计{table_name}失败: {e}")
                        ws.append([period, '查询失败', str(e)])

                # 添加股票活跃度统计（如果有amount字段）
                if self._check_table_column_exists(primary_table, 'amount'):
                    try:
                        activity_sql = f"""
                        SELECT
                            stock_code,
                            AVG(amount) as avg_amount,
                            COUNT(*) as trading_days
                        FROM {primary_table}
                        WHERE period = '{primary_period}'
                          AND amount > 0
                          AND trade_date >= DATE_SUB(CURDATE(), INTERVAL 30 DAY)
                        GROUP BY stock_code
                        ORDER BY avg_amount DESC
                        LIMIT 10
                        """

                        activity_stats = self.db_manager.query_to_dataframe(activity_sql)

                        if not activity_stats.empty:
                            ws.append(["", ""])
                            ws.append(["活跃股票TOP10 (近30天平均成交金额)", ""])
                            ws.append(["股票代码", "平均成交金额(元)", "交易天数"])

                            for _, row in activity_stats.iterrows():
                                ws.append([
                                    row['stock_code'],
                                    f"{row['avg_amount']:,.2f}",
                                    int(row['trading_days'])
                                ])

                    except Exception as e:
                        logger.warning(f"活跃度统计失败: {e}")

                logger.info("市场统计信息导出完成")
            else:
                ws.append(["统计项", "数值"])
                ws.append(["数据查询失败", ""])
                logger.warning("市场统计数据查询失败")

        except Exception as e:
            logger.error(f"导出市场统计失败: {e}")
            # 创建错误信息页面
            try:
                ws = wb.create_sheet("市场统计")
                ws.append(["错误信息", str(e)])
                ws.append(["建议", "1. 运行 python fix_database.py 修复"])
                ws.append(["建议", "2. 检查basic_data_{period}表是否存在"])
                ws.append(["建议", "3. 确保数据库连接正常"])
            except:
                pass

    def _check_table_exists(self, table_name):
        """检查表是否存在"""
        try:
            sql = f"SHOW TABLES LIKE '{table_name}'"
            result = self.db_manager.query_to_dataframe(sql)
            return not result.empty
        except:
            return False

    def _export_tick_data_sample(self, wb: Workbook):
        """导出分笔数据样本"""
        try:
            ws = wb.create_sheet("分笔数据样本")
            # 获取最新的tick_data表和数据
            tick_stats = self._get_tick_data_statistics()

            if tick_stats['latest_tick_date'] is None:
                ws['A1'] = "暂无分笔数据"
                return

            # 查找最新的表
            tables_sql = "SHOW TABLES LIKE 'tick_data_%'"
            tables_result = enhanced_db_manager.query_to_dataframe(tables_sql)

            if tables_result.empty:
                ws['A1'] = "暂无分笔数据表"
                return

            # 找到最新的表
            table_names = []
            for _, row in tables_result.iterrows():
                table_name = list(row.values)[0]
                if 'tick_data_' in table_name:
                    table_names.append(table_name)

            if not table_names:
                ws['A1'] = "暂无有效的分笔数据表"
                return

            table_names.sort(reverse=True)
            latest_table = table_names[0]

            # 获取最新日期的分笔数据样本
            sql = f"""
            SELECT t.stock_code, s.stock_name, t.trade_time, t.price,
                   t.volume, t.amount, t.trade_type
            FROM {latest_table} t
            LEFT JOIN stock_info s ON t.stock_code = s.stock_code
            WHERE t.trade_date = (
                SELECT MAX(trade_date) FROM {latest_table}
            )
            ORDER BY t.trade_time DESC
            LIMIT 10000
            """
            df = enhanced_db_manager.query_to_dataframe(sql)
            if not df.empty:
                # 重命名列
                df = df.rename(columns={
                    'stock_code': '股票代码',
                    'stock_name': '股票名称',
                    'trade_time': '交易时间',
                    'price': '成交价',
                    'volume': '成交量',
                    'amount': '成交额',
                    'trade_type': '交易类型'
                })
                # 添加标题
                ws['A1'] = f"分笔数据样本 (来自{latest_table}, 最新交易日前10000条)"
                ws.merge_cells('A1:G1')
                ws['A1'].style = self.title_style
                # 添加数据
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)
                # 应用样式
                self._apply_table_style(ws, len(df) + 1, 7, start_row=2)
                # 自动调整列宽
                self._auto_fit_columns(ws)
                logger.info(f"分笔数据样本导出完成: {len(df)} 条记录")
            else:
                ws['A1'] = "暂无分笔数据"

        except Exception as e:
            logger.error(f"导出分笔数据样本失败: {e}")

    def _export_tick_data_summary(self, wb):
        """导出分笔数据摘要"""
        try:
            ws = wb.create_sheet("分笔数据摘要")

            # 获取最新的分笔数据表
            today = datetime.now().strftime('%Y%m%d')
            yesterday = (datetime.now() - timedelta(days=1)).strftime('%Y%m%d')

            # 尝试多个可能的表名
            possible_tables = [f"tick_data_{today}", f"tick_data_{yesterday}"]
            table_name = None

            for table in possible_tables:
                check_table_sql = f"SHOW TABLES LIKE '{table}'"
                table_exists = self.db_manager.query_to_dataframe(check_table_sql)
                if not table_exists.empty:
                    table_name = table
                    break

            if not table_name:
                ws.append(["分笔数据", "表不存在", f"未找到今日或昨日的分笔数据表"])
                logger.warning("未找到可用的分笔数据表")
                return

            # 在查询前设置连接字符集
            try:
                charset_sql = "SET NAMES utf8mb4 COLLATE utf8mb4_unicode_ci"
                self.db_manager.execute_sql(charset_sql)

                # 彻底修复表的字符集
                fix_tick_charset_sql = f"ALTER TABLE {table_name} CONVERT TO CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci"
                self.db_manager.execute_sql(fix_tick_charset_sql)

                fix_stock_charset_sql = "ALTER TABLE stock_info CONVERT TO CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci"
                self.db_manager.execute_sql(fix_stock_charset_sql)

            except Exception as charset_error:
                logger.warning(f"修复字符集失败: {charset_error}")

            # 使用分步查询避免字符集冲突
            # 第一步：获取分笔数据
            tick_sql = f"""
            SELECT stock_code, trade_time, price, volume, amount, trade_type
            FROM {table_name}
            WHERE trade_date = (
                SELECT MAX(trade_date) FROM {table_name}
                WHERE trade_date IS NOT NULL
            )
            ORDER BY trade_time DESC
            LIMIT 1000
            """

            df = self.db_manager.query_to_dataframe(tick_sql)

            if df.empty:
                # 如果按trade_date查询为空，尝试其他方式
                tick_sql_alt = f"""
                SELECT stock_code, trade_time, price, volume, amount, trade_type
                FROM {table_name}
                ORDER BY trade_time DESC
                LIMIT 1000
                """
                df = self.db_manager.query_to_dataframe(tick_sql_alt)

            if df.empty:
                ws.append(["股票代码", "股票名称", "交易时间", "价格", "成交量", "成交金额", "交易类型"])
                ws.append(["暂无数据", "", "", "", "", "", ""])
                logger.info("分笔数据为空")
                return

            # 第二步：分别获取股票名称，避免JOIN冲突
            stock_names = {}
            unique_codes = df['stock_code'].unique()

            for code in unique_codes[:50]:  # 限制查询数量避免性能问题
                try:
                    # 使用参数化查询避免字符集问题
                    name_sql = "SELECT stock_name FROM stock_info WHERE stock_code = %s"
                    name_result = self.db_manager.execute_query_with_params(name_sql, (code,))

                    if name_result and len(name_result) > 0:
                        stock_names[code] = name_result[0][0] if name_result[0][0] else '未知'
                    else:
                        stock_names[code] = '未知'

                except Exception as e:
                    logger.warning(f"获取股票 {code} 名称失败: {e}")
                    stock_names[code] = '未知'

            # 添加股票名称到DataFrame
            df['stock_name'] = df['stock_code'].map(stock_names).fillna('未知')

            # 重新排列列顺序
            df = df[['stock_code', 'stock_name', 'trade_time', 'price', 'volume', 'amount', 'trade_type']]

            # 添加标题
            ws['A1'] = f"分笔数据摘要 (来自{table_name})"
            ws.merge_cells('A1:G1')
            ws['A1'].style = self.title_style

            # 添加列标题
            headers = ['股票代码', '股票名称', '交易时间', '价格', '成交量', '成交金额', '交易类型']
            ws.append(headers)

            # 添加数据
            for _, row in df.iterrows():
                ws.append([
                    row['stock_code'],
                    row['stock_name'],
                    str(row['trade_time']),
                    float(row['price']) if row['price'] else 0,
                    int(row['volume']) if row['volume'] else 0,
                    float(row['amount']) if row['amount'] else 0,
                    str(row['trade_type']) if row['trade_type'] else ''
                ])

            # 应用样式
            self._apply_table_style(ws, len(df) + 2, 7, start_row=2)

            # 自动调整列宽
            self._auto_fit_columns(ws)

            logger.info(f"分笔数据摘要导出完成: {len(df)} 条记录")

        except Exception as e:
            logger.error(f"导出分笔数据摘要失败: {e}")
            # 创建错误信息页面
            try:
                ws = wb.create_sheet("分笔数据摘要")
                ws.append(["错误信息", str(e)])
                ws.append(["建议", "1. 检查数据库字符集设置"])
                ws.append(["建议", "2. 运行 python fix_database.py 修复"])
                ws.append(["建议", "3. 确保分笔数据表存在"])
            except:
                pass

    def _apply_table_style(self, ws, row_count: int, col_count: int, start_row: int = 1):
        """应用表格样式"""
        try:
            # 应用表头样式
            for col in range(1, col_count + 1):
                cell = ws.cell(row=start_row, column=col)
                cell.style = self.header_style

            # 应用数据样式
            for row in range(start_row + 1, row_count + start_row):
                for col in range(1, col_count + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.style = self.data_style

        except Exception as e:
            logger.error(f"应用表格样式失败: {e}")

    def _auto_fit_columns(self, ws):
        """自动调整列宽"""
        try:
            for column in ws.columns:
                max_length = 0
                column_letter = None

                for cell in column:
                    # 跳过合并单元格
                    if hasattr(cell, 'coordinate') and cell.coordinate in ws.merged_cells:
                        continue

                    # 获取列字母
                    if column_letter is None:
                        if hasattr(cell, 'column_letter'):
                            column_letter = cell.column_letter
                        else:
                            # 处理MergedCell对象没有column_letter属性的情况
                            from openpyxl.utils import get_column_letter
                            column_letter = get_column_letter(cell.column)

                    try:
                        cell_value = str(cell.value) if cell.value is not None else ""
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                    except:
                        pass

                if column_letter:
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

        except Exception as e:
            logger.error(f"自动调整列宽失败: {e}")

    def _add_conditional_formatting(self, ws, column: str, row_count: int):
        """添加条件格式"""
        try:
            # 为涨跌幅列添加颜色条件格式
            red_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            green_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")

            range_address = f"{column}3:{column}{row_count}"

            # 添加数据条
            data_bar = DataBarRule(
                start_type='num', start_value=-10,
                end_type='num', end_value=10,
                color="4CAF50"
            )
            ws.conditional_formatting.add(range_address, data_bar)

        except Exception as e:
            logger.error(f"添加条件格式失败: {e}")

    def export_stock_detail_by_code(self, stock_code: str, days: int = 90) -> str:
        """导出单个股票的详细数据"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = self.output_dir / f"股票_{stock_code}_详细数据_{timestamp}.xlsx"

            wb = Workbook()
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # 获取股票基本信息
            stock_info_sql = "SELECT * FROM stock_info WHERE stock_code = :stock_code"
            stock_info = enhanced_db_manager.query_to_dataframe(stock_info_sql, {'stock_code': stock_code})

            if stock_info.empty:
                raise ValueError(f"股票代码 {stock_code} 不存在")

            stock_name = stock_info.iloc[0]['stock_name']

            # 1. 基本信息sheet
            info_ws = wb.create_sheet(f"{stock_name}_基本信息")
            for r in dataframe_to_rows(stock_info, index=False, header=True):
                info_ws.append(r)
            self._auto_fit_columns(info_ws)

            # 2. K线数据sheet
            basic_sql = f"""
            SELECT * FROM basic_data
            WHERE stock_code = :stock_code AND period = 'daily'
              AND trade_date >= DATE_SUB(CURDATE(), INTERVAL {days} DAY)
            ORDER BY trade_date DESC
            """
            basic_data = enhanced_db_manager.query_to_dataframe(basic_sql, {'stock_code': stock_code})

            if not basic_data.empty:
                basic_ws = wb.create_sheet(f"{stock_name}_K线数据")
                for r in dataframe_to_rows(basic_data, index=False, header=True):
                    basic_ws.append(r)
                self._auto_fit_columns(basic_ws)

            # 3. 技术指标sheet
            indicator_sql = f"""
            SELECT * FROM indicator_data
            WHERE stock_code = :stock_code
              AND trade_date >= DATE_SUB(CURDATE(), INTERVAL {days} DAY)
            ORDER BY trade_date DESC, indicator_name
            """
            indicator_data = enhanced_db_manager.query_to_dataframe(indicator_sql, {'stock_code': stock_code})

            if not indicator_data.empty:
                indicator_ws = wb.create_sheet(f"{stock_name}_技术指标")
                for r in dataframe_to_rows(indicator_data, index=False, header=True):
                    indicator_ws.append(r)
                self._auto_fit_columns(indicator_ws)

            wb.save(filename)
            logger.info(f"股票 {stock_code} 详细数据导出完成: {filename}")
            return str(filename)

        except Exception as e:
            logger.error(f"导出股票 {stock_code} 详细数据失败: {e}")
            raise


# 创建全局实例
enhanced_excel_exporter = EnhancedExcelExporter()